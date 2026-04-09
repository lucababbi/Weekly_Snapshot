from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.cell.cell import TIME_TYPES  # For date detection
from dateutil.relativedelta import relativedelta
from openpyxl.drawing.image import Image
from openpyxl_image_loader import SheetImageLoader
import polars as pl

# Update the Excel Dashboard
def openwb(loc):
    return load_workbook(f"{loc}/Excel_Dashboard.xlsx", data_only=False)

def writewb(wb, df, sheet_name, start_row=1, start_col=1, clear_sheet=False, 
           autofit=True, hide_gridlines=True, date_col="Date"):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    if clear_sheet:
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None

    # Headers
    for i, col in enumerate(df.columns, start=start_col):
        ws.cell(row=start_row, column=i, value=col)

    # Data (Polars dates auto-convert)
    data = df.to_dicts()
    for r_idx, row in enumerate(data, start=start_row + 1):
        for c_idx, col in enumerate(df.columns, start=start_col):
            cell = ws.cell(row=r_idx, column=c_idx)
            value = row.get(col)
            cell.value = value
            # **FORMAT DATE COLUMN**
            if col.lower() == date_col.lower() and isinstance(value, TIME_TYPES):
                cell.number_format = 'dd/mm/yyyy'

    if hide_gridlines:
        ws.sheet_view.showGridLines = False

    if autofit:
        autofit_worksheet(ws)

def savewb(wb, loc):
    wb.save(f"{loc}/Excel_Dashboard.xlsx")
    wb.close()

def Update_Excel(loc, Time_Series, FX_Series, Returns_Data, FX_Returns, Index_List, EDate, Sector):

    wb = openwb(loc)

    # Update main Tabs
    writewb(wb, Time_Series, "Time_Series_Indices", clear_sheet=True)
    writewb(wb, FX_Series, "FX_Series", clear_sheet=True)
    writewb(wb, Returns_Data, "Annual_Returns_Indices", clear_sheet=True)
    writewb(wb, FX_Returns, "Annual_Returns_FX", clear_sheet=True)
    writewb(wb, Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=True), "Sector_Returns", clear_sheet=True)

    # Update support Tabs
    writewb(wb, Time_Series.select(pl.col(["Date", ".GDAXI", ".STOXX50E", ".STOXX", ".STXWAP"])).filter(pl.col("Date") >= EDate - relativedelta(months=1)).with_columns([(1000 * pl.col(col) / pl.col(col).first()).alias(col) for col in [col for col in Time_Series.select(pl.col(["Date", ".GDAXI", ".STOXX50E", ".STOXX", ".STXWAP"])).columns if col != "Date"]]), "1M_Time_Series_Indices", clear_sheet=True)
    writewb(wb, Time_Series.select(pl.col(["Date", ".V2TX", ".V1XI"])).filter(pl.col("Date") >= EDate - relativedelta(months=12)), "1Y_Time_Series_Indices", clear_sheet=True)
    writewb(wb, FX_Series.filter(pl.col("Date") >= EDate - relativedelta(months=1)).with_columns([(100 * pl.col(col) / pl.col(col).first()).alias(col) for col in [col for col in FX_Series.columns if col != "Date"]]), "1M_Time_Series_FX", clear_sheet=True)

    # Create the Worksheet object for the Dashboard
    Dashboard_WS = wb["Dashboard"]
    Image_Loader = SheetImageLoader(Dashboard_WS)

    # Update specific Cells
    Dashboard_WS["C7"] = f"All as of Close " + EDate.strftime("%d/%m/%d") # Update Date in Dashboard
    Dashboard_WS["D8"] = Time_Series.select(pl.col(["Date", ".STOXX50E"])).sort("Date")[".STOXX50E"].tail(1)[0] # EURO STOXX 50
    Dashboard_WS["D9"] = Time_Series.select(pl.col(["Date", ".STOXX50"])).sort("Date")[".STOXX50"].tail(1)[0] # STOXX Europe 50
    Dashboard_WS["D10"] = Time_Series.select(pl.col(["Date", ".STOXX"])).sort("Date")[".STOXX"].tail(1)[0] # STOXX Europe 600
    Dashboard_WS["D11"] = Time_Series.select(pl.col(["Date", ".SX50UP"])).sort("Date")[".SX50UP"].tail(1)[0] # STOXX USA 500
    Dashboard_WS["D12"] = Time_Series.select(pl.col(["Date", ".SX50UL"])).sort("Date")[".SX50UL"].tail(1)[0] # STOXX USA 500 USD Price
    Dashboard_WS["D13"] = Time_Series.select(pl.col(["Date", ".SXP1E"])).sort("Date")[".SXP1E"].tail(1)[0] # STOXX Asia/Pacific 600
    Dashboard_WS["D14"] = Time_Series.select(pl.col(["Date", ".STXWAP"])).sort("Date")[".STXWAP"].tail(1)[0] # STOXX World AC Universal All Cap

    Dashboard_WS["D18"] = Time_Series.select(pl.col(["Date", ".GDAXI"])).sort("Date")[".GDAXI"].tail(1)[0] # DAX
    Dashboard_WS["D19"] = Time_Series.select(pl.col(["Date", ".MDAXI"])).sort("Date")[".MDAXI"].tail(1)[0] # MDAX
    Dashboard_WS["D20"] = Time_Series.select(pl.col(["Date", ".SDAXI"])).sort("Date")[".SDAXI"].tail(1)[0] # SDAX

    Dashboard_WS["D67"] = Time_Series.select(pl.col(["Date", ".V2TX"])).sort("Date")[".V2TX"].tail(1)[0] # EURO STOXX 50 Volatility (VSTOXX)
    Dashboard_WS["D68"] = Time_Series.select(pl.col(["Date", ".V1XI"])).sort("Date")[".V1XI"].tail(1)[0] # VDAX

    Dashboard_WS["D85"] = FX_Series.select(pl.col(["Date", "EUR17H="])).sort("Date")["EUR17H="].tail(1)[0] # EURUSD
    Dashboard_WS["D86"] = FX_Series.select(pl.col(["Date", "EURCHF17H="])).sort("Date")["EURCHF17H="].tail(1)[0] # EURCHF
    Dashboard_WS["D87"] = FX_Series.select(pl.col(["Date", "EURGBP17H="])).sort("Date")["EURGBP17H="].tail(1)[0] # EURGBP

    ###################
    # Sector Top 3 #
    ###################

    # Spot the Top 3 Sector Indices
    Sector_Top3 = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=True).select(pl.col("Instrument")).head(3)
    Sector_Bottom3 = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=True).select(pl.col("Instrument")).tail(3)

    # Index Name Top 3
    Dashboard_WS["C44"] = Index_List.filter(pl.col("RIC") == Sector_Top3.row(0)[0]).select(pl.col("Full_Name")).row(0)[0].replace("STOXX Europe 600 ", "")
    Dashboard_WS["C45"] = Index_List.filter(pl.col("RIC") == Sector_Top3.row(1)[0]).select(pl.col("Full_Name")).row(0)[0].replace("STOXX Europe 600 ", "")
    Dashboard_WS["C46"] = Index_List.filter(pl.col("RIC") == Sector_Top3.row(2)[0]).select(pl.col("Full_Name")).row(0)[0].replace("STOXX Europe 600 ", "")

    # Close Top 3
    Dashboard_WS["D44"] = Time_Series.select(pl.col(["Date", Sector_Top3.row(0)[0]])).tail(1).get_column(Sector_Top3.row(0)[0])[0]
    Dashboard_WS["D45"] = Time_Series.select(pl.col(["Date", Sector_Top3.row(1)[0]])).tail(1).get_column(Sector_Top3.row(1)[0])[0]
    Dashboard_WS["D46"] = Time_Series.select(pl.col(["Date", Sector_Top3.row(2)[0]])).tail(1).get_column(Sector_Top3.row(2)[0])[0]

    # WoW
    Dashboard_WS["E44"] = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=True).select(pl.col("WoW")).row(0)[0] * 100
    Dashboard_WS["E45"] = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=True).select(pl.col("WoW")).row(1)[0] * 100
    Dashboard_WS["E46"] = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=True).select(pl.col("WoW")).row(2)[0] * 100

    # 1M
    Dashboard_WS["F44"] = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=True).select(pl.col("1 Month")).row(0)[0] * 100
    Dashboard_WS["F45"] = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=True).select(pl.col("1 Month")).row(1)[0] * 100
    Dashboard_WS["F46"] = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=True).select(pl.col("1 Month")).row(2)[0] * 100

    # YTD
    Dashboard_WS["G44"] = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=True).select(pl.col("YTD")).row(0)[0] * 100
    Dashboard_WS["G45"] = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=True).select(pl.col("YTD")).row(1)[0] * 100
    Dashboard_WS["G46"] = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=True).select(pl.col("YTD")).row(2)[0] * 100

    # 1 Year
    Dashboard_WS["H44"] = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=True).select(pl.col("1 Year")).row(0)[0] * 100
    Dashboard_WS["H45"] = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=True).select(pl.col("1 Year")).row(1)[0] * 100
    Dashboard_WS["H46"] = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=True).select(pl.col("1 Year")).row(2)[0] * 100

    # 3 Year
    Dashboard_WS["I44"] = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=True).select(pl.col("3 Year")).row(0)[0] * 100
    Dashboard_WS["I45"] = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=True).select(pl.col("3 Year")).row(1)[0] * 100
    Dashboard_WS["I46"] = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=True).select(pl.col("3 Year")).row(2)[0] * 100

    ###################
    # Sector Bottom 3 #
    ###################

    # Close Bottom 3
    Dashboard_WS["D48"] = Time_Series.select(pl.col(["Date", Sector_Bottom3.row(0)[0]])).tail(1).get_column(Sector_Bottom3.row(0)[0])[0]
    Dashboard_WS["D49"] = Time_Series.select(pl.col(["Date", Sector_Bottom3.row(1)[0]])).tail(1).get_column(Sector_Bottom3.row(1)[0])[0]
    Dashboard_WS["D50"] = Time_Series.select(pl.col(["Date", Sector_Bottom3.row(2)[0]])).tail(1).get_column(Sector_Bottom3.row(2)[0])[0]

    # Index Name Top 3
    Dashboard_WS["C48"] = Index_List.filter(pl.col("RIC") == Sector_Bottom3.row(0)[0]).select(pl.col("Full_Name")).row(0)[0].replace("STOXX Europe 600 ", "")
    Dashboard_WS["C49"] = Index_List.filter(pl.col("RIC") == Sector_Bottom3.row(1)[0]).select(pl.col("Full_Name")).row(0)[0].replace("STOXX Europe 600 ", "")
    Dashboard_WS["C50"] = Index_List.filter(pl.col("RIC") == Sector_Bottom3.row(2)[0]).select(pl.col("Full_Name")).row(0)[0].replace("STOXX Europe 600 ", "")

    # WoW
    Dashboard_WS["E48"] = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=False).select(pl.col("WoW")).row(2)[0] * 100
    Dashboard_WS["E49"] = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=False).select(pl.col("WoW")).row(1)[0] * 100
    Dashboard_WS["E50"] = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=False).select(pl.col("WoW")).row(0)[0] * 100

    # 1M
    Dashboard_WS["F48"] = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=False).select(pl.col("1 Month")).row(2)[0] * 100
    Dashboard_WS["F49"] = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=False).select(pl.col("1 Month")).row(1)[0] * 100
    Dashboard_WS["F50"] = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=False).select(pl.col("1 Month")).row(0)[0] * 100

    # YTD
    Dashboard_WS["G48"] = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=False).select(pl.col("YTD")).row(2)[0] * 100
    Dashboard_WS["G49"] = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=False).select(pl.col("YTD")).row(1)[0] * 100
    Dashboard_WS["G50"] = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=False).select(pl.col("YTD")).row(0)[0] * 100

    # 1 Year
    Dashboard_WS["H48"] = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=False).select(pl.col("1 Year")).row(2)[0] * 100
    Dashboard_WS["H49"] = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=False).select(pl.col("1 Year")).row(1)[0] * 100
    Dashboard_WS["H50"] = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=False).select(pl.col("1 Year")).row(0)[0] * 100

    # 3 Year
    Dashboard_WS["I48"] = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=False).select(pl.col("3 Year")).row(2)[0] * 100
    Dashboard_WS["I49"] = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=False).select(pl.col("3 Year")).row(1)[0] * 100
    Dashboard_WS["I50"] = Returns_Data.filter(pl.col("Instrument").is_in(Sector)).sort("WoW", descending=False).select(pl.col("3 Year")).row(0)[0] * 100

   # FX Performance
   # WoW
    Dashboard_WS["E85"] = FX_Returns.filter(pl.col("Instrument") == "EUR17H=").select(pl.col("WoW")).row(0)[0] * 100
    Dashboard_WS["E86"] = FX_Returns.filter(pl.col("Instrument") == "EURCHF17H=").select(pl.col("WoW")).row(0)[0] * 100
    Dashboard_WS["E87"] = FX_Returns.filter(pl.col("Instrument") == "EURGBP17H=").select(pl.col("WoW")).row(0)[0] * 100

    # 1M
    Dashboard_WS["F85"] = FX_Returns.filter(pl.col("Instrument") == "EUR17H=").select(pl.col("1 Month")).row(0)[0] * 100
    Dashboard_WS["F86"] = FX_Returns.filter(pl.col("Instrument") == "EURCHF17H=").select(pl.col("1 Month")).row(0)[0] * 100
    Dashboard_WS["F87"] = FX_Returns.filter(pl.col("Instrument") == "EURGBP17H=").select(pl.col("1 Month")).row(0)[0] * 100

    # YTD
    Dashboard_WS["G85"] = FX_Returns.filter(pl.col("Instrument") == "EUR17H=").select(pl.col("YTD")).row(0)[0] * 100
    Dashboard_WS["G86"] = FX_Returns.filter(pl.col("Instrument") == "EURCHF17H=").select(pl.col("YTD")).row(0)[0] * 100
    Dashboard_WS["G87"] = FX_Returns.filter(pl.col("Instrument") == "EURGBP17H=").select(pl.col("YTD")).row(0)[0] * 100

    # 1 Year
    Dashboard_WS["H85"] = FX_Returns.filter(pl.col("Instrument") == "EUR17H=").select(pl.col("1 Year")).row(0)[0] * 100
    Dashboard_WS["H86"] = FX_Returns.filter(pl.col("Instrument") == "EURCHF17H=").select(pl.col("1 Year")).row(0)[0] * 100
    Dashboard_WS["H87"] = FX_Returns.filter(pl.col("Instrument") == "EURGBP17H=").select(pl.col("1 Year")).row(0)[0] * 100

    # 3 Year
    Dashboard_WS["I85"] = FX_Returns.filter(pl.col("Instrument") == "EUR17H=").select(pl.col("3 Year")).row(0)[0] * 100
    Dashboard_WS["I86"] = FX_Returns.filter(pl.col("Instrument") == "EURCHF17H=").select(pl.col("3 Year")).row(0)[0] * 100
    Dashboard_WS["I87"] = FX_Returns.filter(pl.col("Instrument") == "EURGBP17H=").select(pl.col("3 Year")).row(0)[0] * 100

    # VIX Performance
    # WoW
    Dashboard_WS["E67"] = Returns_Data.filter(pl.col("Instrument") == ".V2TX").select(pl.col("WoW")).row(0)[0] * 100
    Dashboard_WS["E68"] = Returns_Data.filter(pl.col("Instrument") == ".V1XI").select(pl.col("WoW")).row(0)[0] * 100

    # 1M
    Dashboard_WS["F67"] = Returns_Data.filter(pl.col("Instrument") == ".V2TX").select(pl.col("1 Month")).row(0)[0] * 100
    Dashboard_WS["F68"] = Returns_Data.filter(pl.col("Instrument") == ".V1XI").select(pl.col("1 Month")).row(0)[0] * 100

    # YTD
    Dashboard_WS["G67"] = Returns_Data.filter(pl.col("Instrument") == ".V2TX").select(pl.col("YTD")).row(0)[0] * 100
    Dashboard_WS["G68"] = Returns_Data.filter(pl.col("Instrument") == ".V1XI").select(pl.col("YTD")).row(0)[0] * 100

    # 1 Year
    Dashboard_WS["H67"] = Returns_Data.filter(pl.col("Instrument") == ".V2TX").select(pl.col("1 Year")).row(0)[0] * 100
    Dashboard_WS["H68"] = Returns_Data.filter(pl.col("Instrument") == ".V1XI").select(pl.col("1 Year")).row(0)[0] * 100

    # 3 Year
    Dashboard_WS["I67"] = Returns_Data.filter(pl.col("Instrument") == ".V2TX").select(pl.col("3 Year")).row(0)[0] * 100
    Dashboard_WS["I68"] = Returns_Data.filter(pl.col("Instrument") == ".V1XI").select(pl.col("3 Year")).row(0)[0] * 100

    # DAX Performance
    # WoW
    Dashboard_WS["E18"] = Returns_Data.filter(pl.col("Instrument") == ".GDAXI").select(pl.col("WoW")).row(0)[0] * 100
    Dashboard_WS["E19"] = Returns_Data.filter(pl.col("Instrument") == ".MDAXI").select(pl.col("WoW")).row(0)[0] * 100
    Dashboard_WS["E20"] = Returns_Data.filter(pl.col("Instrument") == ".SDAXI").select(pl.col("WoW")).row(0)[0] * 100

    # 1M
    Dashboard_WS["F18"] = Returns_Data.filter(pl.col("Instrument") == ".GDAXI").select(pl.col("1 Month")).row(0)[0] * 100
    Dashboard_WS["F19"] = Returns_Data.filter(pl.col("Instrument") == ".MDAXI").select(pl.col("1 Month")).row(0)[0] * 100
    Dashboard_WS["F20"] = Returns_Data.filter(pl.col("Instrument") == ".SDAXI").select(pl.col("1 Month")).row(0)[0] * 100

    # YTD
    Dashboard_WS["G18"] = Returns_Data.filter(pl.col("Instrument") == ".GDAXI").select(pl.col("YTD")).row(0)[0] * 100
    Dashboard_WS["G19"] = Returns_Data.filter(pl.col("Instrument") == ".MDAXI").select(pl.col("YTD")).row(0)[0] * 100
    Dashboard_WS["G20"] = Returns_Data.filter(pl.col("Instrument") == ".SDAXI").select(pl.col("YTD")).row(0)[0] * 100

    # 1 Year
    Dashboard_WS["H18"] = Returns_Data.filter(pl.col("Instrument") == ".GDAXI").select(pl.col("1 Year")).row(0)[0] * 100
    Dashboard_WS["H19"] = Returns_Data.filter(pl.col("Instrument") == ".MDAXI").select(pl.col("1 Year")).row(0)[0] * 100
    Dashboard_WS["H20"] = Returns_Data.filter(pl.col("Instrument") == ".SDAXI").select(pl.col("1 Year")).row(0)[0] * 100

    # 3 Year
    Dashboard_WS["I18"] = Returns_Data.filter(pl.col("Instrument") == ".GDAXI").select(pl.col("3 Year")).row(0)[0] * 100
    Dashboard_WS["I19"] = Returns_Data.filter(pl.col("Instrument") == ".MDAXI").select(pl.col("3 Year")).row(0)[0] * 100
    Dashboard_WS["I20"] = Returns_Data.filter(pl.col("Instrument") == ".SDAXI").select(pl.col("3 Year")).row(0)[0] * 100

    # STOXX Indices Performance
    # WoW
    Dashboard_WS["E8"] = Returns_Data.filter(pl.col("Instrument") == ".STOXX50E").select(pl.col("WoW")).row(0)[0] * 100 # EURO STOXX 50
    Dashboard_WS["E9"] = Returns_Data.filter(pl.col("Instrument") == ".STOXX50").select(pl.col("WoW")).row(0)[0] * 100 # STOXX Europe 50
    Dashboard_WS["E10"] = Returns_Data.filter(pl.col("Instrument") == ".STOXX").select(pl.col("WoW")).row(0)[0] * 100 # STOXX Europe 600
    Dashboard_WS["E11"] = Returns_Data.filter(pl.col("Instrument") == ".SX50UP").select(pl.col("WoW")).row(0)[0] * 100 # STOXX USA 500
    Dashboard_WS["E12"] = Returns_Data.filter(pl.col("Instrument") == ".SX50UL").select(pl.col("WoW")).row(0)[0] * 100 # STOXX USA 500 USD Price
    Dashboard_WS["E13"] = Returns_Data.filter(pl.col("Instrument") == ".SXP1E").select(pl.col("WoW")).row(0)[0] * 100 # STOXX Asia/Pacific 600
    Dashboard_WS["E14"] = Returns_Data.filter(pl.col("Instrument") == ".STXWAP").select(pl.col("WoW")).row(0)[0] * 100 # STOXX World AC Universal All Cap

    # 1M
    Dashboard_WS["F8"] = Returns_Data.filter(pl.col("Instrument") == ".STOXX50E").select(pl.col("1 Month")).row(0)[0] * 100 # EURO STOXX 50
    Dashboard_WS["F9"] = Returns_Data.filter(pl.col("Instrument") == ".STOXX50").select(pl.col("1 Month")).row(0)[0] * 100 # STOXX Europe 50
    Dashboard_WS["F10"] = Returns_Data.filter(pl.col("Instrument") == ".STOXX").select(pl.col("1 Month")).row(0)[0] * 100 # STOXX Europe 600
    Dashboard_WS["F11"] = Returns_Data.filter(pl.col("Instrument") == ".SX50UP").select(pl.col("1 Month")).row(0)[0] * 100 # STOXX USA 500
    Dashboard_WS["F12"] = Returns_Data.filter(pl.col("Instrument") == ".SX50UL").select(pl.col("1 Month")).row(0)[0] * 100 # STOXX USA 500 USD Price
    Dashboard_WS["F13"] = Returns_Data.filter(pl.col("Instrument") == ".SXP1E").select(pl.col("1 Month")).row(0)[0] * 100 # STOXX Asia/Pacific 600
    Dashboard_WS["F14"] = Returns_Data.filter(pl.col("Instrument") == ".STXWAP").select(pl.col("1 Month")).row(0)[0] * 100 # STOXX World AC Universal All Cap

    # YTD
    Dashboard_WS["G8"] = Returns_Data.filter(pl.col("Instrument") == ".STOXX50E").select(pl.col("YTD")).row(0)[0] * 100 # EURO STOXX 50
    Dashboard_WS["G9"] = Returns_Data.filter(pl.col("Instrument") == ".STOXX50").select(pl.col("YTD")).row(0)[0] * 100 # STOXX Europe 50
    Dashboard_WS["G10"] = Returns_Data.filter(pl.col("Instrument") == ".STOXX").select(pl.col("YTD")).row(0)[0] * 100 # STOXX Europe 600
    Dashboard_WS["G11"] = Returns_Data.filter(pl.col("Instrument") == ".SX50UP").select(pl.col("YTD")).row(0)[0] * 100 # STOXX USA 500
    Dashboard_WS["G12"] = Returns_Data.filter(pl.col("Instrument") == ".SX50UL").select(pl.col("YTD")).row(0)[0] * 100 # STOXX USA 500 USD Price
    Dashboard_WS["G13"] = Returns_Data.filter(pl.col("Instrument") == ".SXP1E").select(pl.col("YTD")).row(0)[0] * 100 # STOXX Asia/Pacific 600
    Dashboard_WS["G14"] = Returns_Data.filter(pl.col("Instrument") == ".STXWAP").select(pl.col("YTD")).row(0)[0] * 100 # STOXX World AC Universal All Cap

    # 1Y
    Dashboard_WS["H8"] = Returns_Data.filter(pl.col("Instrument") == ".STOXX50E").select(pl.col("1 Year")).row(0)[0] * 100 # EURO STOXX 50
    Dashboard_WS["H9"] = Returns_Data.filter(pl.col("Instrument") == ".STOXX50").select(pl.col("1 Year")).row(0)[0] * 100 # STOXX Europe 50
    Dashboard_WS["H10"] = Returns_Data.filter(pl.col("Instrument") == ".STOXX").select(pl.col("1 Year")).row(0)[0] * 100 # STOXX Europe 600
    Dashboard_WS["H11"] = Returns_Data.filter(pl.col("Instrument") == ".SX50UP").select(pl.col("1 Year")).row(0)[0] * 100 # STOXX USA 500
    Dashboard_WS["H12"] = Returns_Data.filter(pl.col("Instrument") == ".SX50UL").select(pl.col("1 Year")).row(0)[0] * 100 # STOXX USA 500 USD Price
    Dashboard_WS["H13"] = Returns_Data.filter(pl.col("Instrument") == ".SXP1E").select(pl.col("1 Year")).row(0)[0] * 100 # STOXX Asia/Pacific 600
    Dashboard_WS["H14"] = Returns_Data.filter(pl.col("Instrument") == ".STXWAP").select(pl.col("1 Year")).row(0)[0] * 100 # STOXX World AC Universal All Cap

    # 3Y
    Dashboard_WS["I8"] = Returns_Data.filter(pl.col("Instrument") == ".STOXX50E").select(pl.col("3 Year")).row(0)[0] * 100 # EURO STOXX 50
    Dashboard_WS["I9"] = Returns_Data.filter(pl.col("Instrument") == ".STOXX50").select(pl.col("3 Year")).row(0)[0] * 100 # STOXX Europe 50
    Dashboard_WS["I10"] = Returns_Data.filter(pl.col("Instrument") == ".STOXX").select(pl.col("3 Year")).row(0)[0] * 100 # STOXX Europe 600
    Dashboard_WS["I11"] = Returns_Data.filter(pl.col("Instrument") == ".SX50UP").select(pl.col("3 Year")).row(0)[0] * 100 # STOXX USA 500
    Dashboard_WS["I12"] = Returns_Data.filter(pl.col("Instrument") == ".SX50UL").select(pl.col("3 Year")).row(0)[0] * 100 # STOXX USA 500 USD Price
    Dashboard_WS["I13"] = Returns_Data.filter(pl.col("Instrument") == ".SXP1E").select(pl.col("3 Year")).row(0)[0] * 100 # STOXX Asia/Pacific 600
    Dashboard_WS["I14"] = Returns_Data.filter(pl.col("Instrument") == ".STXWAP").select(pl.col("3 Year")).row(0)[0] * 100 # STOXX World AC Universal All Cap

    # Fill the top row
    Dashboard_WS["C2"].fill = PatternFill(
        start_color="FFA9D2F6",
        end_color="FFA9D2F6",
        fill_type="solid"
    )

    Dashboard_WS["C2"] = "📈 Weekly Benchmarks Snapshot 📉"
    Dashboard_WS["C2"].font = Font(color="FF2A4676", bold=True, size=14)
    Dashboard_WS["C2"].alignment = Alignment(horizontal="center", vertical="center")

    Dashboard_WS.row_dimensions[2].height = 5

    # Autofit remaining columns
    for col in ["D"]:
        max_len = 0
        for cell in Dashboard_WS[col]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        Dashboard_WS.column_dimensions[col].width = min(max(max_len + 2, 10), 40)  # Add padding, set min/max width

    savewb(wb, loc)

def autofit_worksheet(ws, min_width=10, max_width=50, padding=2):
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)

        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))

        ws.column_dimensions[col_letter].width = min(max(max_len + padding, min_width), max_width)

    